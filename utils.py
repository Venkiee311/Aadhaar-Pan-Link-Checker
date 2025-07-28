import logging
import re
from typing import List, Optional

import openpyxl
from pydantic import BaseModel, ValidationError, field_validator

logging.basicConfig(
    level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s"
)

class Employee(BaseModel):
    empno: int
    employee_name: str
    pan_no: str
    aadhar_no: str

    @field_validator("pan_no", "aadhar_no")
    @classmethod
    def sanitize_identifiers(cls, v: str) -> str:
        if v is None:
            raise ValueError("PAN/Aadhaar cannot be None")
        return re.sub(r"\s+", "", str(v))

def validate_excel_file(file_path: str) -> bool:
    try:
        workbook = openpyxl.load_workbook(file_path)
        return True
    except Exception as e:
        logging.error(f"Invalid Excel file: {e}")
        return False

def read_excel_data(file_path: str, sheet_name: str) -> List[Employee]:
    employees = []
    try:
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook[sheet_name]
        header = [cell.value for cell in sheet[1]]

        logging.info(f"Found headers: {header}")

        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue

            raw_data = dict(zip(header, row))
            logging.debug(f"Raw data: {raw_data}")

            mapped_data = {}

            for key, value in raw_data.items():
                if key is None or value is None:
                    continue

                key_clean = str(key).strip()

                if key_clean == "EMPNO":
                    mapped_data["empno"] = value
                elif key_clean == "EMPLOYEE NAME":
                    mapped_data["employee_name"] = value
                elif key_clean == "PAN NO":
                    mapped_data["pan_no"] = value
                elif key_clean == "Aadhar NO":
                    mapped_data["aadhar_no"] = value

            required_fields = ["empno", "employee_name", "pan_no", "aadhar_no"]
            if all(
                field in mapped_data and mapped_data[field] is not None
                for field in required_fields
            ):
                try:
                    empno = mapped_data["empno"]
                    employee_name = mapped_data["employee_name"]
                    pan_no = mapped_data["pan_no"]
                    aadhar_no = mapped_data["aadhar_no"]

                    if (
                        empno is not None
                        and employee_name is not None
                        and pan_no is not None
                        and aadhar_no is not None
                    ):
                        employees.append(
                            Employee(
                                empno=empno,
                                employee_name=employee_name,
                                pan_no=pan_no,
                                aadhar_no=aadhar_no,
                            )
                        )
                except ValidationError as e:
                    logging.error(f"Data validation error for row {row}: {e}")
            else:
                missing_fields = [
                    f
                    for f in required_fields
                    if f not in mapped_data or mapped_data[f] is None
                ]
                logging.warning(
                    f"Incomplete data for row {row}, missing: {missing_fields}"
                )

    except FileNotFoundError:
        logging.error(f"Input Excel file not found at: {file_path}")
    except KeyError:
        logging.error(f"Sheet '{sheet_name}' not found in the Excel file.")
    return employees